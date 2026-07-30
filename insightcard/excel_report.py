"""
Excel (.xlsx) exporter — a multi-sheet workbook with the same numbers
as the PDF, but editable: a small business owner can filter, sort, or
build their own pivot on top of this instead of being stuck with a
static report. Uses openpyxl's native chart objects (not embedded
images) so charts stay editable/resizable in Excel too.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from insightcard.config import BrandConfig
from insightcard.metrics import ReportMetrics

_HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_WARN_FONT = Font(color="C62828")
_OK_FONT = Font(color="2E7D32")

# How many slices/bars a native Excel chart shows before it becomes
# unreadable -- found via a stress test with 40 categories rendering
# as an unreadable pie chart. The full data table above each chart is
# never limited -- only the chart itself.
_MAX_PIE_SLICES = 8
_MAX_BAR_ITEMS = 12


def _style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_series_sheet(wb: Workbook, title: str, index_label: str, series, value_label: str = "Revenue"):
    ws = wb.create_sheet(title)
    ws.append([index_label, value_label])
    _style_header_row(ws, 1, 2)
    for idx, value in series.items():
        ws.append([str(idx), float(value)])
    _autosize_columns(ws, [28, 16])
    return ws


def build_workbook(
    metrics: ReportMetrics,
    brand: BrandConfig,
    output_path: str | Path,
    data_quality=None,
) -> Path:
    wb = Workbook()

    # ------------------------------------------------------------ Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = brand.report_title
    ws["A1"].font = Font(size=16, bold=True, color="1A237E")
    ws["A2"] = brand.business_name
    ws["A2"].font = Font(size=12, italic=True)
    if metrics.date_range:
        ws["A3"] = f"Period: {metrics.date_range[0]} to {metrics.date_range[1]}"

    row = 5
    kpi_rows = [
        ("Total Revenue", f"{brand.currency_symbol}{metrics.total_revenue:,.2f}"),
        ("Orders", f"{metrics.order_count:,}"),
        ("Average Order Value", f"{brand.currency_symbol}{metrics.average_order_value:,.2f}"),
    ]
    if metrics.growth_pct is not None:
        kpi_rows.append(("Growth (last period)", f"{metrics.growth_pct:+.1f}%"))
    for label, value in kpi_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    if data_quality is not None:
        ws.cell(row=row, column=1, value="Data Health Check").font = Font(size=13, bold=True, color="1A237E")
        row += 1
        if data_quality.is_clean:
            cell = ws.cell(row=row, column=1, value=f"No data quality issues detected across {data_quality.total_rows:,} rows.")
            cell.font = _OK_FONT
            row += 1
        else:
            for warning in data_quality.warnings:
                cell = ws.cell(row=row, column=1, value=f"- {warning}")
                cell.font = _WARN_FONT
                row += 1
        row += 1

    ws.cell(row=row, column=1, value="Key Insights").font = Font(size=13, bold=True, color="1A237E")
    row += 1
    for insight in metrics.insights:
        ws.cell(row=row, column=1, value=f"- {insight}")
        row += 1
    _autosize_columns(ws, [90])

    # -------------------------------------------------------- Top Products
    if metrics.top_products is not None and len(metrics.top_products):
        ws2 = _write_series_sheet(wb, "Top Products", "Product", metrics.top_products)
        chart = BarChart()
        chart.title = "Top Products by Revenue"
        chart.y_axis.title = "Revenue"
        data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws2.add_chart(chart, "D2")

    # ---------------------------------------------------------- Monthly
    if metrics.monthly_revenue is not None and len(metrics.monthly_revenue) >= 2:
        monthly_named = metrics.monthly_revenue.copy()
        monthly_named.index = monthly_named.index.astype(str)
        ws3 = _write_series_sheet(wb, "Monthly Revenue", "Month", monthly_named)
        chart = LineChart()
        chart.title = "Monthly Revenue Trend"
        chart.y_axis.title = "Revenue"
        data = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws3.add_chart(chart, "D2")

    # ------------------------------------------------------------ Category
    if metrics.top_categories is not None and len(metrics.top_categories) >= 2:
        ws4 = _write_series_sheet(wb, "Category Breakdown", "Category", metrics.top_categories)
        # Chart only the top rows -- a pie with dozens of slivers is
        # unreadable. The full breakdown is still in the sheet above;
        # this only limits what the CHART displays.
        chart_max_row = min(ws4.max_row, 1 + _MAX_PIE_SLICES)
        chart = PieChart()
        chart.title = f"Revenue Share by Category (top {chart_max_row - 1})" if ws4.max_row > chart_max_row else "Revenue Share by Category"
        data = Reference(ws4, min_col=2, min_row=1, max_row=chart_max_row)
        cats = Reference(ws4, min_col=1, min_row=2, max_row=chart_max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws4.add_chart(chart, "D2")

    # ------------------------------------------------------------- Region
    if metrics.top_regions is not None and len(metrics.top_regions) >= 2:
        ws5 = _write_series_sheet(wb, "Region Breakdown", "Region", metrics.top_regions)
        chart_max_row = min(ws5.max_row, 1 + _MAX_BAR_ITEMS)
        chart = BarChart()
        chart.title = f"Revenue by Region (top {chart_max_row - 1})" if ws5.max_row > chart_max_row else "Revenue by Region"
        chart.y_axis.title = "Revenue"
        data = Reference(ws5, min_col=2, min_row=1, max_row=chart_max_row)
        cats = Reference(ws5, min_col=1, min_row=2, max_row=chart_max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws5.add_chart(chart, "D2")

    # --------------------------------------------------------------- Day
    if metrics.weekday_revenue is not None and len(metrics.weekday_revenue) >= 2:
        _write_series_sheet(wb, "Day of Week", "Day", metrics.weekday_revenue)

    # ------------------------------------------------------- Slow Moving
    if metrics.slow_moving_products is not None and len(metrics.slow_moving_products):
        ws6 = wb.create_sheet("Slow-Moving Stock")
        ws6.append(["Product", "Days Since Last Sale"])
        _style_header_row(ws6, 1, 2)
        for name, days in metrics.slow_moving_products.items():
            ws6.append([str(name), int(days)])
        _autosize_columns(ws6, [28, 22])

    output_path = Path(output_path)
    wb.save(str(output_path))
    return output_path
